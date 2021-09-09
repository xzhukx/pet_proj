from openpyxl import Workbook, load_workbook

class new:
    def __init__(self, a):
        self.a = a
        self.dict = None

    def iter(self):
        _dict = {}
        for i in self.a:
            _dict.update({i:i})
        self.dict =_dict

    def save(self, file_n):
        if self.dict:
            file = open(file_n, "w")
            file.write(str(self.dict))
            file.close()

    def read(self, file_n):
        file = open(file_n, "r")
        self.dict = eval(file.read())
        file.close()

    def save_xls(self):
        dict = self.dict
        wb = Workbook()
        ws = wb.create_sheet("new_dict", 0)
        ws["A1"] = "key"
        ws["B1"] = "value"
        for row, (key, value) in enumerate(dict.items(), start=2):
            ws[f"A{row}"] = key
            ws[f"B{row}"] = value
        wb.save("new.xlsx")
        wb.close()

obj=new(a=[3,11,23,2,3123,123,12,312,31,1])
obj.iter()
obj.save("new")
obj.read("new")
obj.save_xls()